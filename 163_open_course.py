# coding=utf-8
import re
import os
import urllib2

import requests
import json
import time
from openpyxl import Workbook, load_workbook

'''
网易公开课有7858个课程28303个视频
https://c.open.163.com/search/search.htm?query=#/search/all
'''


# 得到7868个course
def get_courses():
    courses = []
    for i in range(2):
        r = requests.get(
            'http://c.open.163.com/search/school.do?callback=&school=&pageNum=' + str(i + 1) + '&pSize=4000')
        courses += json.loads(r.text.strip()[1:-1], encoding='utf-8')['result']['dtos']
    import codecs
    json.dump(courses, codecs.open('courses.json', 'w', 'utf-8'), indent=4, ensure_ascii=False)


def get_courses_from_dwr():
    dwr_file = open('courses.dwr' + '.txt', 'a')
    for i in range(1):
        print i + 1
        data = {
            'callCount': 1,
            'scriptSessionId': '',
            'httpSessionId': '',
            'c0-scriptName': 'OpenSearchBean',
            'c0-methodName': 'searchCourse',
            'c0-id': 0,
            'c0-param0': 'string:',
            'c0-param1': 'number:' + str(i + 1),  # 第一页
            'c0-param2': 'number:3000',  # 3000条每页，超过5000取不到数据
            'batchId': '1493367775649'
        }
        r = requests.post('https://c.open.163.com/dwr/call/plaincall/OpenSearchBean.searchCourse.dwr', data=data)
        # 删除前四行和最后一行
        dwr_file.write('\n'.join(r.text.split('\n')[4:-2]))
    dwr_file.close()


def get_special_from_dwr():
    dwr_file = open('special.dwr' + '.txt', 'a')
    for i in range(1):
        print i + 1
        data = {
            'callCount': 1,
            'scriptSessionId': '',
            'httpSessionId': '',
            'c0-scriptName': 'OpenSearchBean',
            'c0-methodName': 'searchSpecial',
            'c0-id': 0,
            'c0-param0': 'string:',
            'c0-param1': 'number:' + str(i + 1),  # 第一页
            'c0-param2': 'number:500',  # 500条每页，超过5000取不到数据
            'batchId': '1493367775649'
        }
        r = requests.post('https://c.open.163.com/dwr/call/plaincall/OpenSearchBean.searchSpecial.dwr', data=data)
        # 删除前四行和最后一行
        dwr_file.write('\n'.join(r.text.split('\n')[4:-2]))
    dwr_file.close()


def get_videos_from_dwr():
    """
    网易的dwr的规则是offset和limit都不能超过5000所以最多只能取9999条数据
    :return:
    """
    dwr_file = open('videos_dwr' + '.txt', 'a')
    for i in range(4):
        print i + 1
        data = {
            'callCount': 1,
            'scriptSessionId': '${scriptSessionId}190',
            'httpSessionId': '',
            'c0-scriptName': 'OpenSearchBean',
            'c0-methodName': 'searchVideo',
            'c0-id': 0,
            'c0-param0': 'string:',
            'c0-param1': 'number:' + str(i + 1),  # 第i+1页
            'c0-param2': 'number:4999',
            'batchId': '1493470611201'
        }
        r = requests.post('https://c.open.163.com/dwr/call/plaincall/OpenSearchBean.searchVideo.dwr', data=data)
        print(r.text.split('\n'))
        # 删除前四行和最后一行
        dwr_file.write('\n'.join(r.text.split('\n')[4:-2]))
    dwr_file.close()


# 处理后（courseId唯一）有5945个课程，存入mongodb中
def process_courses_dwr():
    courses = []
    dwr_file = open('courses.dwr' + '.txt')
    for line in dwr_file:
        data = {}
        bigPicUrl = re.search(r'bigPicUrl="(.*?)"', line)
        if bigPicUrl:
            data['cover'] = bigPicUrl.group(1).strip()
        category = re.search(r'category="(.*?)"', line)
        if category:
            data['category'] = category.group(1).strip().decode('unicode-escape').encode('utf-8')
        courseId = re.search(r'courseId="(.*?)"', line)
        if courseId:
            data['courseId'] = courseId.group(1).strip()
        courseUrl = re.search(r'courseUrl="(.*?)"', line)
        if courseUrl:
            data['url'] = courseUrl.group(1).strip()
        description = re.search(r'description="(.*?)"', line)
        if description:
            data['description'] = description.group(1).strip().decode('unicode-escape').encode('utf-8')
        instructor = re.search(r'instructor="(.*?)"', line)
        if instructor:
            data['instructor'] = instructor.group(1).strip().decode('unicode-escape').encode('utf-8')
        movieCount = re.search(r'movieCount=(\d+)', line)
        if movieCount:
            data['totalMovieCount'] = int(movieCount.group(1))
        school = re.search(r'school="(.*?)"', line)
        if school:
            data['school'] = school.group(1).strip().decode('unicode-escape').encode('utf-8')
        subject = re.search(r'subject="(.*?)"', line)
        if subject:
            data['subject'] = subject.group(1).strip().decode('unicode-escape').encode('utf-8')
        tags = re.search(r'tags="(.*?)"', line)
        if tags:
            data['tags'] = tags.group(1).strip().decode('unicode-escape').encode('utf-8')
        title = re.search(r'title="(.*?)"', line)
        if title:
            data['title'] = title.group(1).strip().decode('unicode-escape').encode('utf-8')
        courses.append(data)
    return courses


def process_special_dwr():
    courses = []
    dwr_file = open('special.dwr' + '.txt')
    for line in dwr_file:
        data = {}
        bigPicUrl = re.search(r'bigPicUrl="(.*?)"', line)
        if bigPicUrl:
            data['cover'] = bigPicUrl.group(1).strip()
        category = re.search(r'category="(.*?)"', line)
        if category:
            data['category'] = category.group(1).strip().decode('unicode-escape').encode('utf-8')
        courseId = re.search(r'courseId="(.*?)"', line)
        if courseId:
            data['courseId'] = courseId.group(1).strip()
        courseUrl = re.search(r'="(.*?)"', line)
        if courseUrl:
            data['url'] = courseUrl.group(1).strip()
        description = re.search(r'description="(.*?)"', line)
        if description:
            try:
                data['description'] = description.group(1).strip().decode('unicode-escape').encode('utf-8')
            except:
                data['description'] = ''
        instructor = re.search(r'instructor="(.*?)"', line)
        if instructor:
            data['instructor'] = instructor.group(1).strip().decode('unicode-escape').encode('utf-8')
        movieCount = re.search(r'movieCount=(\d+)', line)
        if movieCount:
            data['totalMovieCount'] = int(movieCount.group(1))
        school = re.search(r'school="(.*?)"', line)
        if school:
            data['school'] = school.group(1).strip().decode('unicode-escape').encode('utf-8')
        subject = re.search(r'subject="(.*?)"', line)
        if subject:
            data['subject'] = subject.group(1).strip().decode('unicode-escape').encode('utf-8')
        tags = re.search(r'tags="(.*?)"', line)
        if tags:
            data['tags'] = tags.group(1).strip().decode('unicode-escape').encode('utf-8')
        title = re.search(r'title="(.*?)"', line)
        if title:
            data['title'] = title.group(1).strip().decode('unicode-escape').encode('utf-8')
        courses.append(data)
    return courses


def process_videos_dwr():
    videos = []
    dwr_file = open('videos_dwr' + '.txt')
    i = 1
    for line in dwr_file:
        data = {}
        picUrl = re.search(r'imgurl=(.*?)_180x100x1x95.jpg', line)
        data['cover'] = picUrl.group(1).strip()
        url = re.search(r'url="(.*?)"', line)
        data['url'] = url.group(1).strip()
        tem = re.search(r'^(.*?)\.html', data['url'].split('/')[-1]).group(1)
        data['courseId'] = tem.split('_')[0]
        data['videoId'] = tem.split('_')[1]
        description = re.search(r'description="(.*?)"', line)
        data['description'] = description.group(1).strip().decode('unicode-escape').encode('utf-8')
        title = re.search(r'title="(.*?)"', line)
        data['title'] = title.group(1).strip().decode('unicode-escape').encode('utf-8')
        print data['title']
        if data in videos:
            #print '存在重复' + str(i)
            #i = i+1
            continue
        videos.append(data)
    return videos


def insert_courses():
    i = 0
    courses = json.load(open('courses.json'), encoding='utf-8')
    from mongoengine.errors import NotUniqueError
    for course in courses:
        for k, v in course.items():
            if type(v) in [str, unicode]:
                if len(v) == 0:
                    del course[k]
                else:
                    course[k] = v.strip()
        if 'courseUrl' in course and 'subject' in course:
            try:
                i += 1
                print i, course['title'].encode('utf-8'), course['tags'].encode('utf-8'), course['category'], course[
                    'instructor'], course['description'].encode('utf-8'), course['subject'].encode('utf-8')

            except NotUniqueError, e:
                print e


def process_data(cate):
    f = open(cate + '.txt')
    xls_file = 'courses.xlsx'
    if os.path.isfile(xls_file):
        wb = load_workbook(filename=xls_file)
    else:
        wb = Workbook()
    ws = wb.create_sheet(title=cate.decode('utf-8'))
    ws.append(
        ['title', 'category', 'cover', 'courseId', 'courseType', 'description', 'instructor', 'movieCount',
         'school', 'startTime', 'subject', 'tags'])
    for l in f:
        a = re.sub(r';s\d+\.', "|**|", l).replace('null', 'None').replace('{##', '').replace('##}', '')
        a = re.sub(r's\d+\.', "", a, 1)
        for i in a.split('|**|'):
            # print i
            exec i
        description = description.decode('unicode-escape').encode('utf-8') if description else None
        instructor = instructor.decode('unicode-escape').encode('utf-8') if instructor else None
        school = school.decode('unicode-escape').encode('utf-8') if school else None
        subject = subject.decode('unicode-escape').encode('utf-8') if subject else None
        tags = tags.decode('unicode-escape').encode('utf-8') if tags else None
        category = category.decode('unicode-escape').encode('utf-8') if category else None
        title = title.decode('unicode-escape').encode('utf-8') if title else None
        courseId = courseId.decode('unicode-escape').encode('utf-8') if title else None
        state = 30
        type = 1
        ws.append(
            [title, category, description, instructor, school,
             subject, tags, courseId, state, type])
    wb.save(xls_file)


# 将课程（云图中的专辑）写入json文件中
def parse_url(courses):
    course2 = []
    i = 1
    from mongoengine.errors import NotUniqueError
    for course in courses:
        for k, v in course.items():
            if type(v) in [str, unicode]:
                if len(v) == 0:
                    del course[k]
                else:
                    course[k] = v.strip()
        if 'url' in course and 'subject' in course:
            try:
                # url = parse_by_url(course['url'])
                # course['url'] = url
                if course['url'] is None:
                    continue
                print i, course['title'], course['url']
                i = i + 1
                course2.append(course)
            except NotUniqueError, e:
                print e
    writeToJson(course2, 'courses_1.json')


def search_video_by_course_id(id, keyword, limit):
    data = {
        'callCount': 1,
        'scriptSessionId': '${scriptSessionId}190',
        'httpSessionId': '',
        'c0-scriptName': 'OpenSearchBean',
        'c0-methodName': 'searchVideo',
        'c0-id': 0,
        'c0-param0': 'string:' + keyword,
        'c0-param1': 'number:1',  # 第一页
        'c0-param2': 'number:' + str(limit),  # 3000条每页，超过3000取不到数据
        'batchId': '1493470611201'
    }
    r = requests.post('https://c.open.163.com/dwr/call/plaincall/OpenSearchBean.searchVideo.dwr', data=data)
    items = '\n'.join(r.text.split('\n')[4:-2])
    if '' != items:
        for item in items.split('\n'):
            videoUrl = re.search(r'url="(.*?)"', item).group(1).strip()
            print(item)
            if id in videoUrl:
                return videoUrl
    return None


# 通过视频的html页面获取视频播放地址
def parse_by_url(url):
    print url
    content = urllib2.urlopen(url, timeout=5)
    reg = r'http://mov.bn.netease.com/.*.m3u8'
    my = re.compile(reg, re.S)
    lists = re.findall(my, content.read())
    content.close()
    if len(lists) > 0:
        url = lists[0]
        if '-list.m3u8' in url:
            url = url.replace('-list.m3u8', '.flv')
        else:
            url = url.replace('.m3u8', '.mp4')
        return url


# 通过视频的html页面获取视频播放地址
def parse_by_url2(url):
    content = urllib2.urlopen(url, timeout=5).read()
    reg = r'http://swf.*.swf'
    my = re.compile(reg, re.S)
    lists = re.findall(my, content)
    if len(lists) > 0:
        url = lists[0]
        if '-list.m3u8' in url:
            url = url.replace('-list.m3u8', '.flv')
        else:
            url = url.replace('.m3u8', '.mp4')
        return url


# 获取video数据
def get_videos():
    get_videos_from_dwr()
    videos = process_videos_dwr()
    return videos


# 获取专辑数据
def get_album():
    get_courses_from_dwr()
    courses = process_courses_dwr()
    parse_url(courses)


# 获取公开课策划--出现问题待解决
def get_open():
    get_special_from_dwr()
    courses = process_special_dwr()
    parse_url(courses)


# 获取公开课数据，视频url转换成MP4结尾的播放地址
def getSource():
    videos2 = []
    videos = get_videos()
    i = 1
    for video in videos:
        try:
            url = parse_by_url(video['url'])
            if url.endswith('list.mp4'):
                url = parse_by_url2(video['url'])
        except:
            continue
        video['url'] = url
        print i, '标题：', video['title'], '视频地址：', url, '专辑id', video['courseId'], '单曲id', video['videoId']
        if video['url'] is None:
            continue
        i = i + 1
        videos2.append(video)
    # 单曲写入json文件
    writeToJson(videos2, 'video.json')


# 将list集合写入json文件的方法
def writeToJson(videos, str):
    import sys
    reload(sys)
    sys.setdefaultencoding('utf-8')
    import codecs
    json.dump(videos, codecs.open(str, 'w', 'utf-8'), indent=4, ensure_ascii=False)


if __name__ == "__main__":
    cates = ['TED', 'BBC', '可汗学院', '国际名校公开课', '中国大学视频公开课', '国立台湾大学公开课']
    get_courses()
    # get_videos_from_dwr()
    # videos = process_videos_dwr()
    # print len(videos)
    # get_open()
    # get_album()
    # getSource()
    # get_courses()
    # courses = json.load(open('courses.json'), encoding='utf-8')
    # parseUrl(courses)
    # insert_courses()
    # parseUrl(courses)
    # get_special_from_dwr()
    # get_videos_from_dwr()
    # search_video_by_course_id("","",100)
    # get_courses()

    # for cate in cates:
    # get_data(cate);
    # process_data(cate)
