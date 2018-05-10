# coding=utf-8
from mongoengine import *
import os
import json
import bson
import codecs
import time
import re
import requests
import scrapy
from elasticsearch import Elasticsearch
from openpyxl import load_workbook, Workbook

"""
http://www.yuntu.io/institution/list?lid=57e10a5ea770286d5ebd4878
http://192.168.0.116:8100/institution/list?lid=57cd06c677c8a33013933975
state:
新创建,待续传 0
待转码 5
转换成功,待审 10
转换失败 9
已补充 15
审核通过 20
审核不通过 19
上架 30
下架 35
删除 40
"""


# noinspection PyUnresolvedReferences
class Library(Document):
    name = StringField()
    node = StringField()
    code = StringField()
    authCode = StringField()
    authLink = StringField()
    character = StringField()
    uploadLink = StringField()
    libraryUrl = StringField()
    clientScope = StringField()
    mirrorLink = StringField()
    ipRanges = StringField()
    badge = StringField()
    areaCode = StringField()
    regionCode = StringField()
    cityCode = StringField()

    def __str__(self):
        return self.name.encode('utf-8')


# noinspection PyUnresolvedReferences
class Extender(Document):
    school = StringField()
    node = StringField()
    number = IntField()

    def __str__(self):
        return self.school.encode('utf-8')


# noinspection PyUnresolvedReferences
class MediaAlbum(Document):
    lid = StringField()
    title = StringField()
    cover = StringField()
    uploadUid = StringField()
    uploadUsername = StringField()
    state = IntField(default=30)
    comment = IntField(default=0)
    tags = ListField(StringField())
    label1 = IntField()
    label2 = IntField()
    label3 = IntField()
    libraryLabel1 = IntField()
    libraryLabel2 = IntField()
    formats = IntField()
    source = StringField()
    favorite = IntField(default=0)
    type = IntField(default=1)  # 1视频 2音频
    timeline = LongField()
    copyrightTime = LongField()
    copyrightOwner = StringField()
    mediaCount = IntField()
    hot = IntField(default=0)
    open = IntField(default=1)
    channel = IntField(default=0)
    view = IntField(default=0)
    isbn = StringField()
    producer = StringField()
    introduction = StringField()
    artist = StringField()
    categoriesName1 = StringField()
    categoriesName2 = StringField()
    category = StringField()
    ethnicity = StringField()
    meta = {'collection': 'mediaAlbum'}

    def __str__(self):
        return self.title.encode('utf-8')


# noinspection PyUnresolvedReferences
class Media(Document):
    lid = StringField()
    aid = StringField()
    timeline = LongField()
    state = IntField()
    label1 = IntField()
    label2 = IntField()
    label3 = IntField()
    title = StringField()
    cover = StringField()
    url = StringField()
    type = IntField()  # 1视频 2音频
    uploadTime = LongField()
    screenshotCount = IntField()
    hot = IntField()
    view = IntField()
    tags = ListField(StringField())
    uploadUsername = StringField()
    uploadUid = StringField()
    open = IntField(default=1)
    introduction = StringField()

    def __str__(self):
        return self.title.encode('utf-8')


# noinspection PyUnresolvedReferences
class CategoriesLibrary(Document):
    """
    机构分类
    """
    lid = StringField()
    name = StringField()
    parent = StringField()
    url = StringField()
    code = IntField()
    order = IntField()
    recommend = IntField()
    idx = IntField()
    meta = {'collection': 'categoriesLibrary'}

    def __str__(self):
        return self.name.encode('utf-8')


class ResourceCategory(Document):
    """
    把资源跟机构和机构分类对应,在机构页面显示
    """
    rid = StringField()
    lid = StringField()
    label1 = IntField()
    label2 = IntField()
    type = IntField()
    meta = {'collection': 'resourceCategory'}


class Nation(Document):
    code = LongField()
    nationName = StringField()
    introduction = StringField()
    chineseName = StringField()
    population = StringField()
    foreignName = StringField()
    distribution = StringField()
    nickName = StringField()
    language = StringField()
    character = StringField()
    backgroundUrl = StringField()


class Course(Document):
    courseId = StringField(unique=True)
    title = StringField()
    neteaseUrl1 = URLField()  # 课程的专属页面
    neteaseUrl2 = URLField()  # 第一个视频的页面
    description = StringField()
    school = StringField()
    source = StringField()
    tags = StringField()
    total = IntField()
    cover = URLField()
    pubDate = StringField()
    instructor = StringField()
    subject = StringField()
    category = StringField()
    aid = StringField()
    videoUrls = ListField(URLField())
    videos = ListField(StringField())
    crawled = BooleanField(default=False)
    relatedCourseIds = ListField(StringField())
    view = IntField()


class Video(Document):
    videoId = StringField()
    docId = StringField()  # 网易公开课用于获取评论
    mid = StringField()
    description = StringField()
    bigPicUrl = URLField()
    title = StringField()
    courseId = StringField()
    url = URLField(unique=True)
    videoUrl = URLField()  # 视频资源地址
    crawled = BooleanField(default=False)
    view = IntField()


class Comment(Document):
    tid = StringField()
    title = StringField()
    type = IntField()  # 目标实体类型：1，视频，2,音频，3，文档，4，图库，5，书本
    typeName = StringField()
    uid = StringField()
    username = StringField()
    timeline = LongField()
    ip = StringField()
    content = StringField()
    score = IntField()
    replyIdList = ListField(StringField())
    praises = IntField()
    replies = IntField()
    url = StringField()  # 评论目标在云图的地址
    courseId = StringField()
    videoId = StringField()


# noinspection PyUnresolvedReferences
class Yuntu(object):
    def __init__(self, JSESSIONID=None):
        self.es = YuntuEs()
        self.client = connect('xfile')
        self.nationFile = 'nations.json'
        self.coursesLogFile = 'courses.log.json'
        self.coursesFile = 'courses.json'
        self.xlsxFile = 'data.xlsx'
        self.videosFile = 'videos.json'
        self.categoriesFile = 'managerCategories.json'
        self.JSESSIONID = JSESSIONID
        self.cookies = {'JSESSIONID': JSESSIONID}
        self.openCourseLid = '57e10a5ea770286d5ebd4878'

    @staticmethod
    def parseTags(tagsString):
        tags = re.split(u'[,;，；、]', tagsString)
        return [tag for tag in tags if tag != '']

    @staticmethod
    def parseCover(cover):
        if not cover or cover == '':
            return None
        if 'oimagec2.ydstatic.com' in cover:
            return re.search(r'url=(.*\.jpg)', cover).group(1)
        elif 'imgsize.ph.126.net' in cover:
            return re.search(r'url=(.*\.jpg)_', cover).group(1)
        else:
            return cover

    @staticmethod
    def parseVideoUrl(videoUrl):
        """
        :type videoUrl: str
        """
        if not videoUrl:
            return None
        if '-list.m3u8' in videoUrl:
            return videoUrl.replace('-list.m3u8', '.flv')
        else:
            return videoUrl.replace('.m3u8', '.mp4')

    def modifyTagsOfCourses(self):
        for course in Course.objects:
            tags = self.parseTags(course.tags) if course.tags else []
            title = course.title
            tags = [tag for tag in tags if
                    tag != course.category and
                    tag != u'网易公开课' and
                    not (tag in title and len(tag) / float(len(unicode(title))) > 0.5)]
            course.update(tags=','.join(tags))

    # def exportNations(self):
    #     wb = load_workbook(filename=self.nationFile)
    #     nationSheet = wb.get_sheet_by_name(u'民族')
    #     for nation in Nation.objects:
    #         nationSheet.append([
    #             nation.code,
    #             nation.chineseName,
    #             nation.foreignName,
    #             nation.nickName,
    #             nation.introduction,
    #             nation.population,
    #             nation.distribution,
    #             nation.language,
    #             nation.character,
    #             nation.backgroundUrl,
    #         ])
    #     wb.save(self.nationFile)
    def exportNations(self):
        nations = []
        for nation in Nation.objects:
            data = json.loads(nation.to_json(), encoding='utf-8')
            del data['_id']
            nations.append(data)
        json.dump(nations, codecs.open(self.nationFile, 'w', encoding='utf-8'), indent=2, ensure_ascii=False)

    def exportCoursesLog(self, course):
        """
        :type course: Course
        """
        data = {
            'courseId': course.courseId,
            'neteaseUrl1': course.neteaseUrl1,
            'neteaseUrl2': course.neteaseUrl2,
            'title': course.title,
        }
        f = codecs.open(self.coursesLogFile, 'a', 'utf-8')
        json.dump(data, f, indent=4, ensure_ascii=False)
        f.write('\n')

    def exportCourse(self, course):
        """
        导出了后crawled字段设置为True
        :type course: Course
        """
        tags = self.parseTags(course.tags) if course.tags else []
        title = course.title
        tags = [tag for tag in tags if
                tag != course.category and
                tag != u'网易公开课' and
                not (tag in title and len(tag) / float(len(unicode(title))) > 0.5)]
        if course.source == u'网易公开课':
            label1 = 230000
        else:
            label1 = None
        if course.category == u'国立台湾大学公开课' or course.category == u'中国大学视频公开课':
            label2 = 230200
        elif course.category == 'TED':
            label2 = 230100
        elif course.category == 'BBC':
            label2 = 230400
        elif course.category == u'国际名校公开课':
            label2 = 230300
        else:
            label2 = 230500
        libraryLabel1, libraryLabel2 = None, None
        if label1 == 230000:
            libraryLabel1 = 12000
            if label2 == 230200:
                libraryLabel2 = 12002
            if label2 == 230300:
                libraryLabel2 = 12004
            if label2 == 230400 or 'BBC' in course.title:
                libraryLabel1 = 12100
                libraryLabel2 = 12102
            if label2 == 230100 or 'TED' in course.title:
                libraryLabel1 = 12200
                libraryLabel2 = 12202
        data = dict(
            id=bson.ObjectId().__str__(),
            title=course.title,
            cover=course.cover,
            timeline=self.getCurrentTimestamp(),
            introduction=course.description,
            tags=tags,
            source=course.source,
            uploadUsername=u'公开课',
            mediaCount=course.videoUrls.__sizeof__(),
            state=30,
            comment=0,
            favorite=0,
            type=1,
            hot=0,
            open=1,
            channel=0,
            view=0,
            artist=course.instructor,
            label1=label1,
            label2=label2,
            libraryLabel1=libraryLabel1,
            libraryLabel2=libraryLabel2,
            courseId=course.courseId,
        )
        course.update(aid=data['id'], crawled=True)
        # if label1 == 230000:
        #     ResourceCategory(rid=data['id'], lid=self.openCourseLid, label1=libraryLabel1, label2=libraryLabel2,
        #                      type=1).save()
        return data

    def exportCourses(self, limit=None):
        courses = []
        count = 0
        for course in Course.objects(crawled=False)[:limit if limit else Course.objects.count()]:
            count += 1
            print count, course.title
            courses.append(self.exportCourse(course))
        json.dump(courses, codecs.open(self.coursesFile, 'w', encoding='utf-8'), indent=2, ensure_ascii=False)

    def exportVideos(self):
        videos = []
        count = 0
        courses = json.load(codecs.open(self.coursesFile, encoding='utf-8'))
        for course in courses:
            for video in Video.objects(courseId=course['courseId']):
                assert isinstance(video, Video)
                data = dict(
                    id=bson.ObjectId().__str__(),
                    title=video.title,
                    url=self.parseVideoUrl(video.videoUrl.__str__()),
                    timeline=self.getCurrentTimestamp(),
                    tags=','.join(course['tags']),
                    cover=self.parseCover(video.bigPicUrl.__str__()),
                    uploadUsername=u'公开课',
                    introduction=video.description,
                    state=30,
                    type=1,
                    open=1,
                    aid=course['id'],
                )
                video.update(mid=data['id'])
                videos.append(data)
                count += 1
                print count
        json.dump(videos, codecs.open(self.videosFile, 'w', encoding='utf-8'), indent=2, ensure_ascii=False)

    def exportNationsXlsx(self):
        nations = json.load(codecs.open(self.nationFile, encoding='utf-8'))
        nationFields = ['code', 'chineseName', 'foreignName', 'nickName', 'introduction', 'population', 'distribution',
                        'language', 'character', 'backgroundUrl']
        if os.path.isfile(self.xlsxFile):
            wb = load_workbook(filename=self.xlsxFile)
            nationSheet = wb.get_sheet_by_name(u'民族')
        else:
            wb = Workbook()
            wb.remove_sheet(wb.get_active_sheet())
            nationSheet = wb.create_sheet(u'民族')
            nationSheet.append(nationFields)
        for nation in nations:
            nationSheet.append([nation[key] for key in nationFields])
        wb.save(self.xlsxFile)

    def exportXlsx(self):
        courses = json.load(codecs.open(self.coursesFile, encoding='utf-8'))
        videos = json.load(codecs.open(self.videosFile, encoding='utf-8'))
        courseFields = ['id', 'title', 'cover', 'label1', 'label2', 'tags', 'introduction', 'uploadUsername', 'artist',
                        'type', 'source', 'libraryLabel1', 'libraryLabel2']
        videoFields = ['id', 'title', 'cover', 'tags', 'introduction', 'uploadUsername', 'artist', 'type', 'source',
                       'url', 'aid']
        if os.path.isfile(self.xlsxFile):
            wb = load_workbook(filename=self.xlsxFile)
            albumSheet = wb.get_sheet_by_name(u'专辑')
            mediaSheet = wb.get_sheet_by_name(u'视频')
        else:
            wb = Workbook()
            wb.remove_sheet(wb.get_active_sheet())
            albumSheet = wb.create_sheet(u'专辑')
            mediaSheet = wb.create_sheet(u'视频')
            albumSheet.append(courseFields)
            mediaSheet.append(videoFields)
        for course in courses:
            course['tags'] = ','.join(course['tags']) if course['tags'] else ''
            albumSheet.append([course[key] for key in courseFields])
        for video in videos:
            video['artist'] = u'公开课'
            video['source'] = u'网易公开课'
            mediaSheet.append([video[key] for key in videoFields])
        wb.save(self.xlsxFile)

    def deleteMediaAlbum(self, mediaAlbum=None, aid=None):
        if mediaAlbum:
            self.es.deleteMediaAlbum(mediaAlbum)
            mediaAlbum.delete()
        elif aid:
            self.es.deleteMediaAlbum(aid=aid)
            MediaAlbum.objects.get(id=aid).delete()
        else:
            # 没有指定专辑，则清除所有专辑
            for mediaAlbum in MediaAlbum.objects:
                print 'delete'
                self.es.deleteMediaAlbum(mediaAlbum)
            MediaAlbum.objects.delete()

    @staticmethod
    def deleteMedia(media=None, mid=None):
        if media:
            media.delete()
        elif mid:
            Media.objects.get(id=mid).delete()
        else:
            Media.objects.delete()  # 没有指定媒体，则清除所有媒体

    @staticmethod
    def getCurrentTimestamp(timeline=long(round(time.time() * 1000))):
        """
        当前时间的13位时间戳
        :rtype: long
        """
        return timeline

    def getManagerLibraries(self):
        r = requests.get('http://manager.libtop.com:9090/library?pageSize=3000', cookies=self.cookies)
        selector = scrapy.Selector(text=r.text)
        items = selector.xpath('//tr/td[1]/a')
        libraries = []
        for item in items:
            libraries.append(dict(
                name=item.xpath('text()').extract_first(),
                id=item.xpath('@href').extract_first().split('/')[-1]
            ))
        json.dump(libraries, codecs.open('managerLibraries.json', 'w', encoding='utf-8'), indent=2, ensure_ascii=False)
        return libraries

    def getManagerExtenders(self):
        r = requests.get('http://manager.libtop.com:9090/extender/list?pageSize=3000', cookies=self.cookies)
        selector = scrapy.Selector(text=r.text)
        items = selector.xpath('//table/tbody/tr')
        extenders = []
        for item in items:
            extenders.append(dict(
                number=item.xpath('td[1]/text()').extract_first(),
                school=item.xpath('td[2]/a/text()').extract_first(),
                id=item.xpath('td[2]/a/@href').extract_first().split('/')[-1]
            ))
        json.dump(extenders, codecs.open('managerExtenders.json', 'w', encoding='utf-8'), indent=2, ensure_ascii=False)
        return extenders

    def getManagerCategories(self):
        r = requests.get('http://manager.libtop.com:9090/categories/tree', cookies=self.cookies)
        selector = scrapy.Selector(text=r.text)
        items = selector.xpath('//*[@id="example1"]/tbody/tr')
        categories = []

        def parseChildren(pId, pCode, pName):
            children = requests.get('http://manager.libtop.com:9090/categories/children.json?pid=' + pId,
                                    cookies=self.cookies).json()
            for child in children:
                print child['name']
                categories.append({'name': child['name'], 'code': child['code'], 'parent': pCode, 'parentName': pName})
                if child['code'] > 990000:
                    parseChildren(child['id'], child['code'], child['name'])

        for item in items:
            parentId = item.xpath('@data-tt-id').extract_first()
            parentName = item.xpath('td[2]/text()').extract_first().strip()
            parentCode = item.xpath('td[1]/text()').extract_first()
            categories.append({'name': parentName, 'code': int(parentCode)})
            parseChildren(parentId, int(parentCode), parentName)
        json.dump(categories, codecs.open(self.categoriesFile, 'w', encoding='utf-8'), indent=2, ensure_ascii=False)
        return categories


class YuntuEs(object):
    def __init__(self):
        self.es = Elasticsearch([{'host': '192.168.0.116', 'port': 9200}])

    def deleteMediaAlbum(self, mediaAlbum=None, aid=None):
        try:
            if mediaAlbum:
                self.es.delete('media-album-index', 'media-album', mediaAlbum.id.__str__())
            if aid:
                self.es.delete('media-album-index', 'media-album', aid)
        except:
            pass

    def saveMediaAlbum(self, mediaAlbum):
        mid = mediaAlbum.id.__str__()
        data = json.loads(mediaAlbum.to_json(), encoding='utf-8')
        data['id'] = mid
        del data['_id']
        self.es.index('media-album-index', 'media-album', data, id=mid)

    def updateMediaAlbum(self, mediaAlbum):
        self.deleteMediaAlbum(mediaAlbum)
        self.saveMediaAlbum(mediaAlbum)


yuntu = Yuntu()
